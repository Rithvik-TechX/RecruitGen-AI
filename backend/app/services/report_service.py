"""
Report Service — generate candidate, hiring, and match reports with PDF export.
"""

from __future__ import annotations

import os
import uuid
from collections import Counter
from datetime import datetime, timezone

import structlog
from fastapi import HTTPException, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.application import Application, ApplicationStatus
from app.models.candidate import CandidateProfile
from app.models.candidate_match import CandidateMatch
from app.models.hiring_recommendation import HiringRecommendation
from app.models.interview_schedule import InterviewSchedule
from app.models.job import Job
from app.models.report import Report, ReportStatus, ReportType
from app.models.resume import Resume
from app.models.user import User
from app.repositories.candidate_match_repository import CandidateMatchRepository
from app.repositories.candidate_ranking_repository import CandidateRankingRepository
from app.repositories.candidate_repository import CandidateProfileRepository
from app.repositories.hiring_recommendation_repository import (
    HiringRecommendationRepository,
)
from app.repositories.report_repository import ReportRepository
from app.repositories.skill_evaluation_repository import SkillEvaluationRepository

logger = structlog.get_logger(__name__)

REPORTS_DIR = os.path.join("uploads", "reports")


class ReportService:
    """Generate and manage recruitment reports."""

    def __init__(self, session: AsyncSession) -> None:
        self._session = session
        self._repo = ReportRepository(session)
        self._candidate_repo = CandidateProfileRepository(session)
        self._match_repo = CandidateMatchRepository(session)
        self._ranking_repo = CandidateRankingRepository(session)
        self._eval_repo = SkillEvaluationRepository(session)
        self._rec_repo = HiringRecommendationRepository(session)

    async def generate_organization_report(
        self,
        report_type: ReportType,
        user: User,
        *,
        title: str | None = None,
    ) -> Report:
        """Generate and persist a presentation-ready organization report."""
        content, summary = await self._build_organization_report_data(report_type, user)
        report = Report(
            generated_by=user.id,
            report_type=report_type,
            title=title or (
                "Organization Hiring Report"
                if report_type == ReportType.HIRING
                else "Organization Analytics Report"
            ),
            summary=summary,
            content=content,
            status=ReportStatus.COMPLETED,
        )
        report = await self._repo.create(report)
        await self._session.commit()
        await self._session.refresh(report)
        logger.info(
            "organization_report_generated",
            report_id=str(report.id),
            report_type=report_type.value,
            organization_id=str(user.organization_id),
        )
        return report

    async def _build_organization_report_data(
        self, report_type: ReportType, user: User,
    ) -> tuple[dict, str]:
        """Build the shared typed presentation model used by UI and exports."""
        jobs = list((
            await self._session.execute(
                select(Job).where(Job.organization_id == user.organization_id)
            )
        ).scalars().all())
        job_ids = [job.id for job in jobs]

        applications = []
        matches = []
        recommendations = []
        interviews = []
        if job_ids:
            applications = list((
                await self._session.execute(
                    select(Application)
                    .where(Application.job_id.in_(job_ids))
                    .order_by(Application.applied_at.asc())
                )
            ).scalars().all())
            matches = list((
                await self._session.execute(
                    select(CandidateMatch).where(CandidateMatch.job_id.in_(job_ids))
                )
            ).scalars().all())
            recommendations = list((
                await self._session.execute(
                    select(HiringRecommendation).where(
                        HiringRecommendation.job_id.in_(job_ids)
                    )
                )
            ).scalars().all())
            interviews = list((
                await self._session.execute(
                    select(InterviewSchedule).where(
                        InterviewSchedule.job_id.in_(job_ids)
                    )
                )
            ).scalars().all())

        status_counts = {
            status_value.value: sum(
                1 for application in applications if application.status == status_value
            )
            for status_value in ApplicationStatus
        }
        total_applications = len(applications)
        screened_count = sum(
            status_counts[value]
            for value in (
                ApplicationStatus.SCREENED.value,
                ApplicationStatus.SHORTLISTED.value,
                ApplicationStatus.INTERVIEW_SCHEDULED.value,
                ApplicationStatus.INTERVIEW_COMPLETED.value,
                ApplicationStatus.SELECTED.value,
            )
        )
        shortlisted_count = sum(
            status_counts[value]
            for value in (
                ApplicationStatus.SHORTLISTED.value,
                ApplicationStatus.INTERVIEW_SCHEDULED.value,
                ApplicationStatus.INTERVIEW_COMPLETED.value,
                ApplicationStatus.SELECTED.value,
            )
        )
        interviewed_count = sum(
            status_counts[value]
            for value in (
                ApplicationStatus.INTERVIEW_SCHEDULED.value,
                ApplicationStatus.INTERVIEW_COMPLETED.value,
                ApplicationStatus.SELECTED.value,
            )
        )

        candidate_user_ids = {application.candidate_id for application in applications}
        profile_rows = []
        if candidate_user_ids:
            profile_rows = (
                await self._session.execute(
                    select(CandidateProfile, Resume.candidate_id)
                    .join(Resume, CandidateProfile.resume_id == Resume.id)
                    .where(Resume.candidate_id.in_(candidate_user_ids))
                    .order_by(CandidateProfile.updated_at.desc())
                )
            ).all()
        profiles_by_user: dict[uuid.UUID, CandidateProfile] = {}
        for profile, candidate_user_id in profile_rows:
            profiles_by_user.setdefault(candidate_user_id, profile)

        match_by_pair = {
            (item.candidate_id, item.job_id): item for item in matches
        }
        recommendation_by_pair = {
            (item.candidate_id, item.job_id): item for item in recommendations
        }
        job_by_id = {job.id: job for job in jobs}
        candidate_insights = []
        applicant_matches = []
        for application in applications:
            profile = profiles_by_user.get(application.candidate_id)
            match = (
                match_by_pair.get((profile.id, application.job_id))
                if profile else None
            )
            recommendation = (
                recommendation_by_pair.get((profile.id, application.job_id))
                if profile else None
            )
            if match:
                applicant_matches.append(match)
            candidate_insights.append({
                "candidate_name": (
                    profile.full_name
                    if profile else application.candidate.full_name
                ),
                "job_applied": job_by_id[application.job_id].title,
                "match_score": self._score_percent(
                    match.overall_match_score if match else None
                ),
                "skill_score": self._score_percent(
                    match.skill_match_score if match else None
                ),
                "status": application.status.value,
                "recommendation": recommendation.decision.value if recommendation else None,
            })

        funnel = {
            "applied": total_applications,
            "screened": screened_count,
            "shortlisted": shortlisted_count,
            "interviewed": interviewed_count,
            "selected": status_counts[ApplicationStatus.SELECTED.value],
            "rejected": status_counts[ApplicationStatus.REJECTED.value],
        }
        trends: dict[str, int] = {}
        for application in applications:
            date_key = application.applied_at.date().isoformat()
            trends[date_key] = trends.get(date_key, 0) + 1

        score_buckets: dict[str, int] = {
            "0-20": 0,
            "21-40": 0,
            "41-60": 0,
            "61-80": 0,
            "81-100": 0,
        }
        for match in applicant_matches:
            percentage = self._score_percent(match.overall_match_score) or 0
            bucket = (
                "0-20" if percentage <= 20 else
                "21-40" if percentage <= 40 else
                "41-60" if percentage <= 60 else
                "61-80" if percentage <= 80 else
                "81-100"
            )
            score_buckets[bucket] += 1

        def conversion(numerator: int, denominator: int) -> float:
            return round(numerator / denominator * 100, 2) if denominator else 0.0

        generated_at = datetime.now(timezone.utc)
        if applications:
            start_date = min(item.applied_at for item in applications).date()
            end_date = max(item.applied_at for item in applications).date()
            report_period = (
                start_date.strftime("%b %d, %Y")
                if start_date == end_date
                else f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
            )
        else:
            report_period = "No applications in this period"

        skill_counts = Counter(
            skill.skill_name
            for profile in profiles_by_user.values()
            for skill in profile.skills
        )
        top_skills = [
            {"skill_name": name, "count": count}
            for name, count in skill_counts.most_common(10)
        ]
        recommendation_summary = {
            "hire": sum(1 for item in recommendations if item.decision.value == "hire"),
            "consider": sum(1 for item in recommendations if item.decision.value == "consider"),
            "reject": sum(1 for item in recommendations if item.decision.value == "reject"),
        }
        header = {
            "organization_name": user.organization.name,
            "generated_date": generated_at.isoformat(),
            "report_period": report_period,
        }
        skills_phrase = ", ".join(item["skill_name"] for item in top_skills[:3])
        executive_summary = (
            f"During this hiring cycle, {total_applications} applications were received "
            f"across {len(jobs)} positions. {screened_count} candidates progressed to "
            f"screening, {interviewed_count} reached interviews, and "
            f"{status_counts[ApplicationStatus.SELECTED.value]} were selected."
        )
        if skills_phrase:
            executive_summary += f" {skills_phrase} were the most common applicant skills."

        if report_type == ReportType.HIRING:
            content = {
                "kind": "hiring",
                "header": header,
                "summary": {
                    "total_jobs": len(jobs),
                    "total_applications": total_applications,
                    "screened": screened_count,
                    "shortlisted": shortlisted_count,
                    "interviewed": interviewed_count,
                    "selected": status_counts[ApplicationStatus.SELECTED.value],
                    "rejected": status_counts[ApplicationStatus.REJECTED.value],
                },
                "pipeline": [
                    {"stage": "Applied", "count": total_applications},
                    {"stage": "Screened", "count": screened_count},
                    {"stage": "Shortlisted", "count": shortlisted_count},
                    {"stage": "Interview", "count": interviewed_count},
                    {
                        "stage": "Selected",
                        "count": status_counts[ApplicationStatus.SELECTED.value],
                    },
                ],
                "candidates": candidate_insights,
                "recommendation_summary": recommendation_summary,
                "top_skills": top_skills,
                "executive_summary": executive_summary,
            }
        else:
            match_scores = [
                self._score_percent(item.overall_match_score) or 0
                for item in applicant_matches
            ]
            skill_scores = [
                self._score_percent(item.skill_match_score) or 0
                for item in applicant_matches
            ]
            content = {
                "kind": "analytics",
                "header": header,
                "summary": {
                    "applications": total_applications,
                    "conversion_rate": conversion(
                        status_counts[ApplicationStatus.SELECTED.value],
                        total_applications,
                    ),
                    "average_match_score": round(
                        sum(match_scores) / len(match_scores), 2
                    ) if match_scores else 0,
                    "average_skill_score": round(
                        sum(skill_scores) / len(skill_scores), 2
                    ) if skill_scores else 0,
                },
                "application_trend": [
                    {"label": date, "value": count}
                    for date, count in sorted(trends.items())
                ],
                "status_distribution": [
                    {"label": key.replace("_", " ").title(), "value": value}
                    for key, value in status_counts.items()
                    if value > 0
                ],
                "recommendation_distribution": [
                    {"label": key.title(), "value": value}
                    for key, value in recommendation_summary.items()
                ],
                "match_score_distribution": [
                    {"label": key, "value": value}
                    for key, value in score_buckets.items()
                ],
                "executive_summary": executive_summary,
            }
        return content, executive_summary

    @staticmethod
    def _score_percent(value: float | None) -> float | None:
        if value is None:
            return None
        return round(value * 100 if value <= 1 else value, 1)

    async def presentation_data(self, report: Report) -> dict:
        """Return typed content, rebuilding legacy JSON reports when necessary."""
        if report.content and report.content.get("kind") in {"hiring", "analytics", "match"}:
            return report.content
        if report.report_type in (ReportType.HIRING, ReportType.ANALYTICS):
            content, _ = await self._build_organization_report_data(
                report.report_type, report.author,
            )
            return content
        if report.report_type == ReportType.MATCH:
            content = report.content or {}
            matches = content.get("matches") or []
            return {
                "kind": "match",
                "header": {
                    "organization_name": report.author.organization.name,
                    "generated_date": report.created_at.isoformat(),
                    "report_period": report.created_at.strftime("%b %d, %Y"),
                },
                "summary": {
                    "total_matches": content.get("total_matches", len(matches)),
                    "average_overall_score": self._score_percent(
                        content.get("avg_overall_score"),
                    ) or 0,
                },
                "matches": [
                    {
                        "candidate": item.get("candidate_name")
                        or item.get("candidate_id")
                        or "Candidate",
                        "overall_score": self._score_percent(item.get("overall_score")) or 0,
                        "skill_score": self._score_percent(item.get("skill_score")) or 0,
                        "experience_score": self._score_percent(
                            item.get("experience_score"),
                        ) or 0,
                        "education_score": self._score_percent(
                            item.get("education_score"),
                        ) or 0,
                        "semantic_score": self._score_percent(
                            item.get("semantic_score"),
                        ) or 0,
                    }
                    for item in matches
                    if isinstance(item, dict)
                ],
                "executive_summary": report.summary or "Match analysis completed.",
            }
        return {
            "kind": "generic",
            "header": {
                "organization_name": report.author.organization.name,
                "generated_date": report.created_at.isoformat(),
                "report_period": report.created_at.strftime("%b %d, %Y"),
            },
            "executive_summary": report.summary or "Report completed.",
        }

    async def generate_candidate_report(
        self, candidate_id: uuid.UUID, user_id: uuid.UUID,
    ) -> Report:
        """Generate a detailed candidate profile report."""
        candidate = await self._candidate_repo.get_by_id(candidate_id)
        if not candidate:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Candidate not found",
            )

        content = {
            "candidate_name": candidate.full_name,
            "email": candidate.email,
            "phone": candidate.phone,
            "linkedin": candidate.linkedin_url,
            "github": candidate.github_url,
            "summary": candidate.summary,
            "skills": [
                {"name": s.skill_name, "level": s.proficiency_level, "years": s.years_of_experience}
                for s in candidate.skills
            ],
            "education": [
                {"institution": e.institution, "degree": e.degree, "field": e.field_of_study}
                for e in candidate.education
            ],
            "experience": [
                {"company": e.company, "title": e.title, "description": e.description}
                for e in candidate.experiences
            ],
            "projects": [
                {"name": p.project_name, "description": p.description}
                for p in candidate.projects
            ],
            "certifications": [
                {"name": c.certification_name, "issuer": c.issuing_organization}
                for c in candidate.certifications
            ],
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

        report = Report(
            generated_by=user_id,
            report_type=ReportType.CANDIDATE,
            title=f"Candidate Report — {candidate.full_name}",
            summary=f"Comprehensive profile report for {candidate.full_name}",
            content=content,
            status=ReportStatus.COMPLETED,
        )
        report = await self._repo.create(report)
        await self._session.commit()
        logger.info("candidate_report_generated", report_id=str(report.id))
        return report

    async def generate_hiring_report(
        self, job_id: uuid.UUID, user_id: uuid.UUID,
    ) -> Report:
        """Generate a hiring summary report for a job."""
        rankings = await self._ranking_repo.list_by_job(job_id)
        matches = await self._match_repo.list_by_job(job_id)
        evaluations = await self._eval_repo.list_by_job(job_id)
        recommendations = await self._rec_repo.list_by_job(job_id)

        content = {
            "job_id": str(job_id),
            "total_candidates_matched": len(matches),
            "total_candidates_ranked": len(rankings),
            "total_evaluated": len(evaluations),
            "total_recommendations": len(recommendations),
            "top_candidates": [
                {
                    "candidate_id": str(r.candidate_id),
                    "rank": r.rank_position,
                    "final_score": r.final_score,
                }
                for r in rankings[:10]
            ],
            "recommendation_summary": {
                "hire": sum(1 for r in recommendations if r.decision.value == "hire"),
                "consider": sum(1 for r in recommendations if r.decision.value == "consider"),
                "reject": sum(1 for r in recommendations if r.decision.value == "reject"),
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

        report = Report(
            job_id=job_id,
            generated_by=user_id,
            report_type=ReportType.HIRING,
            title=f"Hiring Report — Job {str(job_id)[:8]}",
            summary=f"Hiring summary with {len(rankings)} candidates ranked",
            content=content,
            status=ReportStatus.COMPLETED,
        )
        report = await self._repo.create(report)
        await self._session.commit()
        logger.info("hiring_report_generated", report_id=str(report.id))
        return report

    async def generate_match_report(
        self, job_id: uuid.UUID, user_id: uuid.UUID,
    ) -> Report:
        """Generate a match analysis report for a job."""
        matches = await self._match_repo.list_by_job(job_id)

        content = {
            "job_id": str(job_id),
            "total_matches": len(matches),
            "matches": [
                {
                    "candidate_id": str(m.candidate_id),
                    "overall_score": m.overall_match_score,
                    "skill_score": m.skill_match_score,
                    "experience_score": m.experience_match_score,
                    "education_score": m.education_match_score,
                    "semantic_score": m.semantic_similarity_score,
                }
                for m in matches
            ],
            "avg_overall_score": (
                sum(m.overall_match_score for m in matches) / len(matches)
                if matches else 0
            ),
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

        report = Report(
            job_id=job_id,
            generated_by=user_id,
            report_type=ReportType.MATCH,
            title=f"Match Report — Job {str(job_id)[:8]}",
            summary=f"Match analysis with {len(matches)} candidates",
            content=content,
            status=ReportStatus.COMPLETED,
        )
        report = await self._repo.create(report)
        await self._session.commit()
        logger.info("match_report_generated", report_id=str(report.id))
        return report

    async def list_reports(
        self, user_id: uuid.UUID, *, skip: int = 0, limit: int = 50,
    ) -> tuple[list[Report], int]:
        """List reports generated by a user."""
        reports = await self._repo.list_by_user(user_id, skip=skip, limit=limit)
        total = await self._repo.count_by_user(user_id)
        return reports, total

    async def get_report(
        self, report_id: uuid.UUID, *, user_id: uuid.UUID | None = None,
    ) -> Report:
        """Get a single report."""
        report = await self._repo.get_by_id(report_id)
        if not report:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Report not found",
            )
        if user_id is not None and report.generated_by != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to view this report",
            )
        return report

    async def export_pdf(
        self, report_id: uuid.UUID, *, user_id: uuid.UUID | None = None,
    ) -> str:
        """Export a branded, paginated report as PDF."""
        report = await self.get_report(report_id, user_id=user_id)
        data = await self.presentation_data(report)

        os.makedirs(REPORTS_DIR, exist_ok=True)
        file_name = f"report_{report_id}.pdf"
        file_path = os.path.join(REPORTS_DIR, file_name)

        try:
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                KeepTogether,
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )

            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=16 * mm,
                leftMargin=16 * mm,
                topMargin=18 * mm,
                bottomMargin=18 * mm,
                title=report.title,
                author="RecruitGen AI",
            )
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name="Brand",
                parent=styles["Title"],
                fontName="Helvetica-Bold",
                fontSize=20,
                leading=24,
                textColor=colors.HexColor("#0f172a"),
                spaceAfter=4,
            ))
            styles.add(ParagraphStyle(
                name="Section",
                parent=styles["Heading2"],
                fontName="Helvetica-Bold",
                fontSize=12,
                textColor=colors.HexColor("#1d4ed8"),
                spaceBefore=12,
                spaceAfter=7,
            ))
            styles.add(ParagraphStyle(
                name="Executive",
                parent=styles["BodyText"],
                fontSize=9.5,
                leading=14,
                textColor=colors.HexColor("#334155"),
                backColor=colors.HexColor("#f8fafc"),
                borderColor=colors.HexColor("#cbd5e1"),
                borderWidth=0.5,
                borderPadding=10,
            ))
            story: list = []
            header = data["header"]
            generated = datetime.fromisoformat(header["generated_date"])
            story.extend([
                Paragraph("RecruitGen AI", ParagraphStyle(
                    "Company", parent=styles["Normal"], alignment=TA_CENTER,
                    fontName="Helvetica-Bold", fontSize=10,
                    textColor=colors.HexColor("#2563eb"),
                )),
                Paragraph(report.title, styles["Brand"]),
                Paragraph(
                    f"{header['organization_name']} &nbsp;&nbsp;|&nbsp;&nbsp; "
                    f"Generated {generated.strftime('%B %d, %Y')} &nbsp;&nbsp;|&nbsp;&nbsp; "
                    f"{header['report_period']}",
                    styles["Normal"],
                ),
                Spacer(1, 12),
                Paragraph("Executive Summary", styles["Section"]),
                Paragraph(data["executive_summary"], styles["Executive"]),
            ])

            def styled_table(rows: list[list], widths: list[float] | None = None) -> Table:
                table = Table(rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("LEADING", (0, 0), (-1, -1), 11),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                        colors.white, colors.HexColor("#f8fafc"),
                    ]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]))
                return table

            if data["kind"] == "hiring":
                summary = data["summary"]
                story.extend([
                    Paragraph("Hiring Summary", styles["Section"]),
                    styled_table([
                        ["Jobs", "Applications", "Screened", "Shortlisted", "Interviewed", "Selected", "Rejected"],
                        [
                            summary["total_jobs"], summary["total_applications"],
                            summary["screened"], summary["shortlisted"],
                            summary["interviewed"], summary["selected"], summary["rejected"],
                        ],
                    ]),
                    Paragraph("Pipeline Funnel", styles["Section"]),
                    styled_table([
                        [item["stage"] for item in data["pipeline"]],
                        [item["count"] for item in data["pipeline"]],
                    ]),
                    Paragraph("Recommendation Summary", styles["Section"]),
                    styled_table([
                        ["Hire", "Consider", "Reject"],
                        [
                            data["recommendation_summary"]["hire"],
                            data["recommendation_summary"]["consider"],
                            data["recommendation_summary"]["reject"],
                        ],
                    ]),
                    Paragraph("Top Applicant Skills", styles["Section"]),
                    styled_table(
                        [["Skill", "Applicants"]] + [
                            [item["skill_name"], item["count"]]
                            for item in data["top_skills"]
                        ],
                        [130 * mm, 38 * mm],
                    ),
                    PageBreak(),
                    Paragraph("Candidate Insights", styles["Section"]),
                    styled_table(
                        [["Candidate", "Job Applied", "Match", "Skill", "Recommendation", "Status"]] + [
                            [
                                item["candidate_name"],
                                item["job_applied"],
                                f"{item['match_score']:.1f}%" if item["match_score"] is not None else "-",
                                f"{item['skill_score']:.1f}%" if item["skill_score"] is not None else "-",
                                (item["recommendation"] or "Pending").title(),
                                item["status"].replace("_", " ").title(),
                            ]
                            for item in data["candidates"]
                        ],
                        [30 * mm, 42 * mm, 18 * mm, 18 * mm, 28 * mm, 32 * mm],
                    ),
                ])
            elif data["kind"] == "analytics":
                summary = data["summary"]
                story.extend([
                    Paragraph("Analytics Summary", styles["Section"]),
                    styled_table([
                        ["Applications", "Conversion Rate", "Average Match", "Average Skill"],
                        [
                            summary["applications"],
                            f"{summary['conversion_rate']:.1f}%",
                            f"{summary['average_match_score']:.1f}%",
                            f"{summary['average_skill_score']:.1f}%",
                        ],
                    ]),
                ])
                for title, key in (
                    ("Application Trend", "application_trend"),
                    ("Status Distribution", "status_distribution"),
                    ("Recommendation Distribution", "recommendation_distribution"),
                    ("Match Score Distribution", "match_score_distribution"),
                ):
                    story.extend([
                        Paragraph(title, styles["Section"]),
                        styled_table(
                            [["Category", "Value"]] + [
                                [item["label"], item["value"]] for item in data[key]
                            ],
                            [130 * mm, 38 * mm],
                        ),
                    ])
            elif data["kind"] == "match":
                summary = data["summary"]
                story.extend([
                    Paragraph("Match Summary", styles["Section"]),
                    styled_table([
                        ["Total Matches", "Average Overall Score"],
                        [
                            summary["total_matches"],
                            f"{summary['average_overall_score']:.1f}%",
                        ],
                    ]),
                    Paragraph("Candidate Match Scores", styles["Section"]),
                    styled_table(
                        [["Candidate", "Overall", "Skill", "Experience", "Education", "Semantic"]] + [
                            [
                                item["candidate"],
                                f"{item['overall_score']:.1f}%",
                                f"{item['skill_score']:.1f}%",
                                f"{item['experience_score']:.1f}%",
                                f"{item['education_score']:.1f}%",
                                f"{item['semantic_score']:.1f}%",
                            ]
                            for item in data["matches"]
                        ],
                        [42 * mm, 24 * mm, 24 * mm, 28 * mm, 28 * mm, 24 * mm],
                    ),
                ])

            def add_page_number(canvas, document) -> None:
                canvas.saveState()
                canvas.setFont("Helvetica", 8)
                canvas.setFillColor(colors.HexColor("#64748b"))
                canvas.drawString(16 * mm, 10 * mm, "RecruitGen AI - Confidential")
                canvas.drawRightString(
                    A4[0] - 16 * mm, 10 * mm, f"Page {document.page}",
                )
                canvas.restoreState()

            doc.build(
                story,
                onFirstPage=add_page_number,
                onLaterPages=add_page_number,
            )

            report.file_path = file_path
            await self._session.flush()
            await self._session.commit()

            logger.info("pdf_exported", report_id=str(report_id), path=file_path)
        except Exception:
            logger.exception("pdf_export_failed", report_id=str(report_id))
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="PDF export failed",
            )

        return file_path

    async def export_excel(
        self, report_id: uuid.UUID, *, user_id: uuid.UUID | None = None,
    ) -> str:
        """Export a styled four-sheet Excel workbook."""
        report = await self.get_report(report_id, user_id=user_id)
        data = await self.presentation_data(report)
        os.makedirs(REPORTS_DIR, exist_ok=True)
        file_path = os.path.join(REPORTS_DIR, f"report_{report_id}.xlsx")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            workbook = Workbook()
            workbook.remove(workbook.active)
            blue = "1E3A8A"
            pale = "EFF6FF"

            def sheet(name: str):
                ws = workbook.create_sheet(name)
                ws.freeze_panes = "A2"
                return ws

            def title_row(ws, values: list[str]) -> None:
                ws.append(values)
                for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor=blue)
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(vertical="center")

            def finish(ws) -> None:
                ws.auto_filter.ref = ws.dimensions
                for column in ws.columns:
                    width = min(
                        max(len(str(cell.value or "")) for cell in column) + 2,
                        48,
                    )
                    ws.column_dimensions[column[0].column_letter].width = width
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

            summary_ws = sheet("Summary")
            title_row(summary_ws, ["RecruitGen AI Report", "Value"])
            header = data["header"]
            summary_ws.append(["Report Title", report.title])
            summary_ws.append(["Organization", header["organization_name"]])
            summary_ws.append(["Generated Date", header["generated_date"]])
            summary_ws.append(["Report Period", header["report_period"]])
            summary_ws.append(["Executive Summary", data["executive_summary"]])
            for key, value in data.get("summary", {}).items():
                summary_ws.append([key.replace("_", " ").title(), value])
            finish(summary_ws)

            candidates_ws = sheet("Candidates")
            title_row(candidates_ws, [
                "Candidate Name", "Job Applied", "Match Score", "Skill Score",
                "Recommendation", "Status",
            ])
            for item in data.get("candidates", []):
                candidates_ws.append([
                    item["candidate_name"], item["job_applied"],
                    item["match_score"], item["skill_score"],
                    item["recommendation"] or "Pending",
                    item["status"].replace("_", " ").title(),
                ])
            finish(candidates_ws)

            recommendations_ws = sheet("Recommendations")
            title_row(recommendations_ws, ["Recommendation", "Count"])
            for key, value in data.get("recommendation_summary", {}).items():
                recommendations_ws.append([key.title(), value])
            finish(recommendations_ws)

            analytics_ws = sheet("Analytics")
            title_row(analytics_ws, ["Section", "Category", "Value"])
            if data["kind"] == "hiring":
                for item in data["pipeline"]:
                    analytics_ws.append(["Pipeline", item["stage"], item["count"]])
                for item in data["top_skills"]:
                    analytics_ws.append(["Top Skills", item["skill_name"], item["count"]])
            elif data["kind"] == "analytics":
                for section, key in (
                    ("Application Trend", "application_trend"),
                    ("Status Distribution", "status_distribution"),
                    ("Recommendation Distribution", "recommendation_distribution"),
                    ("Match Score Distribution", "match_score_distribution"),
                ):
                    for item in data[key]:
                        analytics_ws.append([section, item["label"], item["value"]])
            elif data["kind"] == "match":
                for item in data.get("matches", []):
                    analytics_ws.append(["Match", item["candidate"], item["overall_score"]])
            finish(analytics_ws)

            for ws in workbook.worksheets:
                ws.sheet_view.showGridLines = False
                for row in ws.iter_rows(min_row=2, max_row=2):
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=pale)
            workbook.save(file_path)
            logger.info("excel_exported", report_id=str(report_id), path=file_path)
        except Exception:
            logger.exception("excel_export_failed", report_id=str(report_id))
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Excel export failed",
            )
        return file_path

    async def export_csv(self, report_id: uuid.UUID, *, user_id: uuid.UUID) -> str:
        """Export report data as CSV."""
        import csv
        import tempfile
        report = await self.get_report(report_id, user_id=user_id)
        data = await self.presentation_data(report)

        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8",
        )
        writer = csv.writer(tmp)

        # Header row
        writer.writerow(["Report", report.title or report.report_type])
        writer.writerow(["Generated", report.created_at.strftime("%Y-%m-%d %H:%M")])
        writer.writerow([])

        summary = data.get("summary", {})
        if summary and isinstance(summary, dict):
            writer.writerow(["Summary"])
            for k, v in summary.items():
                if not any(x in k.lower() for x in ["id", "uuid"]):
                    writer.writerow([k.replace("_", " ").title(), v])
            writer.writerow([])

        if data["kind"] == "match":
            writer.writerow([
                "Candidate", "Overall Score", "Skill Score", "Experience Score",
                "Education Score", "Semantic Score",
            ])
            for item in data.get("matches", []):
                writer.writerow([
                    item["candidate"],
                    item["overall_score"],
                    item["skill_score"],
                    item["experience_score"],
                    item["education_score"],
                    item["semantic_score"],
                ])
        else:
            candidates = data.get("candidates", [])
            if candidates and isinstance(candidates, list):
                headers = [
                    "candidate_name",
                    "job_applied",
                    "match_score",
                    "skill_score",
                    "recommendation",
                    "status",
                ]
                writer.writerow([
                    "Candidate Name", "Job Applied", "Match Score", "Skill Score",
                    "Recommendation", "Status",
                ])
                for row in candidates:
                    if isinstance(row, dict):
                        writer.writerow([row.get(h, "") for h in headers])

        if data.get("executive_summary"):
            writer.writerow([])
            writer.writerow(["Executive Summary", data["executive_summary"]])

        tmp.close()
        return tmp.name
